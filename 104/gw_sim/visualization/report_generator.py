import numpy as np
import os
import logging
from datetime import datetime
from typing import Dict, List, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


class HydrologyReportGenerator:

    HEADER_FONT = Font(name="Arial", size=12, bold=True, color="FFFFFF")
    HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    SUBHEADER_FONT = Font(name="Arial", size=10, bold=True, color="1F3864")
    SUBHEADER_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    DATA_FONT = Font(name="Arial", size=9)
    THIN_BORDER = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    TITLE_FONT = Font(name="Arial", size=16, bold=True, color="1F3864")
    SECTION_FONT = Font(name="Arial", size=11, bold=True, color="2F5496")

    def __init__(self):
        self._wb: Optional[Workbook] = None

    def _create_workbook(self):
        self._wb = Workbook()

    def _apply_header_style(self, ws, row: int, cols: int):
        for col in range(1, cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = self.THIN_BORDER

    def _apply_data_style(self, ws, start_row: int, end_row: int, cols: int):
        for row in range(start_row, end_row + 1):
            for col in range(1, cols + 1):
                cell = ws.cell(row=row, column=col)
                cell.font = self.DATA_FONT
                cell.border = self.THIN_BORDER
                cell.alignment = Alignment(horizontal="center", vertical="center")

    def _write_title(self, ws, title: str, row: int = 1) -> int:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        cell = ws.cell(row=row, column=1, value=title)
        cell.font = self.TITLE_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        return row + 2

    def _write_section_title(self, ws, title: str, row: int) -> int:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        cell = ws.cell(row=row, column=1, value=title)
        cell.font = self.SECTION_FONT
        return row + 1

    def _write_info_table(self, ws, info: Dict, row: int) -> int:
        for key, value in info.items():
            ws.cell(row=row, column=1, value=key).font = self.SUBHEADER_FONT
            ws.cell(row=row, column=1).fill = self.SUBHEADER_FILL
            ws.cell(row=row, column=1).border = self.THIN_BORDER
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
            ws.cell(row=row, column=2, value=str(value)).font = self.DATA_FONT
            ws.cell(row=row, column=2).border = self.THIN_BORDER
            row += 1
        return row + 1

    def generate_simulation_report(
        self,
        simulation_result: Dict,
        output_path: str,
        report_title: str = "Groundwater Hydrology Simulation Report",
        metadata: Optional[Dict] = None,
    ) -> str:
        self._create_workbook()

        ws_summary = self._wb.active
        ws_summary.title = "Summary"

        row = self._write_title(ws_summary, report_title)
        row = self._write_section_title(ws_summary, "Report Information", row)

        report_info = {
            "Generation Time": datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S UTC"),
            "Backend": simulation_result.get("backend", "unknown"),
            "Mode": simulation_result.get("mode", "N/A"),
        }
        if metadata:
            report_info.update(metadata)
        row = self._write_info_table(ws_summary, report_info, row)

        row = self._write_section_title(ws_summary, "Simulation Results Summary", row)

        h = None
        if "h" in simulation_result:
            h = np.array(simulation_result["h"])
        elif "h_final" in simulation_result:
            h = np.array(simulation_result["h_final"])

        if h is not None and h.size > 0:
            stats = {
                "Head Mean (m)": f"{np.mean(h):.4f}",
                "Head Min (m)": f"{np.min(h):.4f}",
                "Head Max (m)": f"{np.max(h):.4f}",
                "Head Std Dev (m)": f"{np.std(h):.4f}",
                "Head Range (m)": f"{np.max(h) - np.min(h):.4f}",
                "Grid Size": f"{h.shape[0]} x {h.shape[1]}" if h.ndim == 2 else str(h.shape),
            }
            if "total_decline" in simulation_result:
                td = np.array(simulation_result["total_decline"])
                stats["Total Decline Mean (m)"] = f"{np.mean(td):.4f}"
                stats["Total Decline Max (m)"] = f"{np.max(td):.4f}"
            row = self._write_info_table(ws_summary, stats, row)

        if "vx" in simulation_result and "vy" in simulation_result:
            vx = np.array(simulation_result["vx"])
            vy = np.array(simulation_result["vy"])
            speed = np.sqrt(vx**2 + vy**2)
            vel_stats = {
                "Velocity Mean (m/s)": f"{np.mean(speed):.6f}",
                "Velocity Max (m/s)": f"{np.max(speed):.6f}",
                "Velocity Min (m/s)": f"{np.min(speed):.6f}",
            }
            row = self._write_section_title(ws_summary, "Seepage Velocity Statistics", row)
            row = self._write_info_table(ws_summary, vel_stats, row)

        if h is not None and h.ndim == 2:
            ws_grid = self._wb.create_sheet("Head Field Data")
            ws_grid.cell(row=1, column=1, value="Row\\Col").font = self.HEADER_FONT
            ws_grid.cell(row=1, column=1).fill = self.HEADER_FILL
            ws_grid.cell(row=1, column=1).border = self.THIN_BORDER
            for i in range(h.shape[1]):
                cell = ws_grid.cell(row=1, column=i + 2, value=f"C{i}")
                cell.font = self.HEADER_FONT
                cell.fill = self.HEADER_FILL
                cell.border = self.THIN_BORDER

            for j in range(h.shape[0]):
                cell = ws_grid.cell(row=j + 2, column=1, value=f"R{j}")
                cell.font = self.SUBHEADER_FONT
                cell.fill = self.SUBHEADER_FILL
                cell.border = self.THIN_BORDER
                for i in range(h.shape[1]):
                    cell = ws_grid.cell(row=j + 2, column=i + 2, value=round(float(h[j, i]), 4))
                    cell.font = self.DATA_FONT
                    cell.border = self.THIN_BORDER
                    cell.number_format = "0.0000"

        if "h_annual" in simulation_result:
            ws_annual = self._wb.create_sheet("Annual Evolution")
            headers = ["Year", "Mean Head (m)", "Min Head (m)", "Max Head (m)", "Std Dev (m)"]
            for i, h_text in enumerate(headers):
                ws_annual.cell(row=1, column=i + 1, value=h_text)
            self._apply_header_style(ws_annual, 1, len(headers))

            h_annual_list = simulation_result["h_annual"]
            if isinstance(h_annual_list, list):
                for year_idx, h_year in enumerate(h_annual_list):
                    h_arr = np.array(h_year)
                    if h_arr.size == 0:
                        continue
                    r = year_idx + 2
                    ws_annual.cell(row=r, column=1, value=year_idx)
                    ws_annual.cell(row=r, column=2, value=round(float(np.mean(h_arr)), 4))
                    ws_annual.cell(row=r, column=3, value=round(float(np.min(h_arr)), 4))
                    ws_annual.cell(row=r, column=4, value=round(float(np.max(h_arr)), 4))
                    ws_annual.cell(row=r, column=5, value=round(float(np.std(h_arr)), 4))
                self._apply_data_style(ws_annual, 2, len(h_annual_list) + 1, len(headers))

        self._wb.save(output_path)
        logger.info(f"Report saved to {output_path}")
        return output_path

    def generate_comparison_report(
        self,
        comparison_data: Dict,
        output_path: str,
        report_title: str = "Multi-Scenario Comparison Report",
    ) -> str:
        self._create_workbook()

        ws = self._wb.active
        ws.title = "Comparison Summary"

        row = self._write_title(ws, report_title)
        row = self._write_section_title(ws, "Scenario Overview", row)

        headers = ["Scenario", "Mean Head (m)", "Min Head (m)", "Max Head (m)",
                   "Std Dev (m)", "Range (m)", "Median (m)", "Q75 (m)"]
        for i, h_text in enumerate(headers):
            ws.cell(row=row, column=i + 1, value=h_text)
        self._apply_header_style(ws, row, len(headers))
        row += 1

        for scenario_name, metrics in comparison_data.get("comparison", {}).items():
            ws.cell(row=row, column=1, value=scenario_name)
            ws.cell(row=row, column=2, value=round(metrics.get("mean", 0), 4))
            ws.cell(row=row, column=3, value=round(metrics.get("min", 0), 4))
            ws.cell(row=row, column=4, value=round(metrics.get("max", 0), 4))
            ws.cell(row=row, column=5, value=round(metrics.get("std", 0), 4))
            ws.cell(row=row, column=6, value=round(metrics.get("range", 0), 4))
            ws.cell(row=row, column=7, value=round(metrics.get("median", 0), 4))
            ws.cell(row=row, column=8, value=round(metrics.get("q75", 0), 4))
            row += 1

        self._apply_data_style(ws, row - len(comparison_data.get("comparison", {})),
                               row - 1, len(headers))

        row += 1
        row = self._write_section_title(ws, "Ranking", row)
        rank_headers = ["Rank", "Scenario", "Metric Value"]
        for i, h_text in enumerate(rank_headers):
            ws.cell(row=row, column=i + 1, value=h_text)
        self._apply_header_style(ws, row, len(rank_headers))
        row += 1

        for metric_name, stats in comparison_data.get("metrics", {}).items():
            best = stats.get("best_scenario")
            worst = stats.get("worst_scenario")
            if best:
                ws.cell(row=row, column=1, value=f"Best {metric_name}")
                ws.cell(row=row, column=2, value=best)
                row += 1
            if worst:
                ws.cell(row=row, column=1, value=f"Worst {metric_name}")
                ws.cell(row=row, column=2, value=worst)
                row += 1

        self._wb.save(output_path)
        logger.info(f"Comparison report saved to {output_path}")
        return output_path

    def generate_long_term_report(
        self,
        projection_result: Dict,
        output_path: str,
        years: int = 10,
        report_title: str = "Long-term Hydrology Projection Report",
    ) -> str:
        self._create_workbook()

        ws = self._wb.active
        ws.title = "Long-term Projection"

        row = self._write_title(ws, report_title)
        row = self._write_section_title(ws, "Projection Parameters", row)

        params_info = {
            "Projection Period": f"{years} years",
            "Generation Time": datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S UTC"),
        }
        h_initial = np.array(projection_result.get("h_initial", []))
        if h_initial.size > 0:
            params_info["Initial Head Mean (m)"] = f"{np.mean(h_initial):.4f}"
        h_final = np.array(projection_result.get("h_final", []))
        if h_final.size > 0:
            params_info["Final Head Mean (m)"] = f"{np.mean(h_final):.4f}"
        td = np.array(projection_result.get("total_decline", []))
        if td.size > 0:
            params_info["Total Decline Mean (m)"] = f"{np.mean(td):.4f}"
            params_info["Total Decline Max (m)"] = f"{np.max(td):.4f}"

        row = self._write_info_table(ws, params_info, row)

        row = self._write_section_title(ws, "Annual Head Statistics", row)
        headers = ["Year", "Mean Head (m)", "Min Head (m)", "Max Head (m)",
                   "Decline from Initial (m)"]
        for i, h_text in enumerate(headers):
            ws.cell(row=row, column=i + 1, value=h_text)
        self._apply_header_style(ws, row, len(headers))
        row += 1

        h_annual = projection_result.get("h_annual", [])
        if isinstance(h_annual, list):
            for year_idx, h_year in enumerate(h_annual):
                h_arr = np.array(h_year)
                if h_arr.size == 0:
                    continue
                decline = float(np.mean(h_initial - h_arr)) if h_initial.size > 0 else 0
                ws.cell(row=row, column=1, value=year_idx)
                ws.cell(row=row, column=2, value=round(float(np.mean(h_arr)), 4))
                ws.cell(row=row, column=3, value=round(float(np.min(h_arr)), 4))
                ws.cell(row=row, column=4, value=round(float(np.max(h_arr)), 4))
                ws.cell(row=row, column=5, value=round(decline, 4))
                row += 1

        for col_idx in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 18

        self._wb.save(output_path)
        logger.info(f"Long-term report saved to {output_path}")
        return output_path
