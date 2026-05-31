from pyspark.sql import SparkSession
from pyspark.sql.functions import (
    col, sum, avg, max, min, count, when, 
    date_format, concat_ws, lit
)
import logging
from datetime import datetime, timedelta
import json
import os
from pathlib import Path

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False
    logger.warning("Pandas not available, some export features may be limited")

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    logger.warning("Openpyxl not available, Excel export may be limited")

class ReportGenerator:
    def __init__(self, spark=None):
        self.spark = spark or SparkSession.builder \
            .appName("ReportGenerator") \
            .config("spark.sql.session.timeZone", "Asia/Shanghai") \
            .enableHiveSupport() \
            .getOrCreate()
        self.report_dir = Path("reports")
        self.report_dir.mkdir(exist_ok=True)

    def generate_daily_report(self, date=None, format='xlsx'):
        if not date:
            date = (datetime.now() - timedelta(days=1)).strftime('%Y-%m-%d')
        
        logger.info(f"Generating daily report for {date}")
        
        report_data = self._collect_report_data(date, date, 'daily')
        filename = f"PV_Daily_Report_{date.replace('-', '')}.{format}"
        
        if format == 'xlsx' and EXCEL_AVAILABLE:
            return self._export_to_excel(report_data, filename, 'daily')
        elif format == 'csv':
            return self._export_to_csv(report_data, filename)
        else:
            return self._export_json(report_data, filename)

    def generate_weekly_report(self, year, week, format='xlsx'):
        logger.info(f"Generating weekly report for {year}W{week}")
        
        start_date = datetime.strptime(f"{year}-W{week}-1", "%Y-W%W-%w").strftime('%Y-%m-%d')
        end_date = datetime.strptime(f"{year}-W{week}-7", "%Y-W%W-%w").strftime('%Y-%m-%d')
        
        report_data = self._collect_report_data(start_date, end_date, 'weekly')
        filename = f"PV_Weekly_Report_{year}W{week:02d}.{format}"
        
        if format == 'xlsx' and EXCEL_AVAILABLE:
            return self._export_to_excel(report_data, filename, 'weekly')
        elif format == 'csv':
            return self._export_to_csv(report_data, filename)
        else:
            return self._export_json(report_data, filename)

    def generate_monthly_report(self, year, month, format='xlsx'):
        logger.info(f"Generating monthly report for {year}-{month}")
        
        start_date = f"{year}-{month:02d}-01"
        if month == 12:
            end_date = f"{year + 1}-01-01"
        else:
            end_date = f"{year}-{month + 1:02d}-01"
        end_date = (datetime.strptime(end_date, '%Y-%m-%d') - timedelta(days=1)).strftime('%Y-%m-%d')
        
        report_data = self._collect_report_data(start_date, end_date, 'monthly')
        filename = f"PV_Monthly_Report_{year}{month:02d}.{format}"
        
        if format == 'xlsx' and EXCEL_AVAILABLE:
            return self._export_to_excel(report_data, filename, 'monthly')
        elif format == 'csv':
            return self._export_to_csv(report_data, filename)
        else:
            return self._export_json(report_data, filename)

    def generate_custom_report(self, start_date, end_date, config=None, format='xlsx'):
        logger.info(f"Generating custom report from {start_date} to {end_date}")
        
        report_data = self._collect_report_data(start_date, end_date, 'custom', config)
        filename = f"PV_Custom_Report_{start_date.replace('-', '')}_{end_date.replace('-', '')}.{format}"
        
        if format == 'xlsx' and EXCEL_AVAILABLE:
            return self._export_to_excel(report_data, filename, 'custom')
        elif format == 'csv':
            return self._export_to_csv(report_data, filename)
        else:
            return self._export_json(report_data, filename)

    def _collect_report_data(self, start_date, end_date, report_type, config=None):
        logger.info("Collecting report data")
        
        data = {
            'report_info': {
                'type': report_type,
                'start_date': start_date,
                'end_date': end_date,
                'generated_at': datetime.now().isoformat()
            },
            'power_summary': self._get_power_summary(start_date, end_date),
            'power_daily': self._get_power_daily(start_date, end_date),
            'fault_summary': self._get_fault_summary(start_date, end_date),
            'fault_distribution': self._get_fault_distribution(start_date, end_date),
            'device_status': self._get_device_status(end_date),
            'efficiency_analysis': self._get_efficiency_analysis(start_date, end_date),
            'loss_analysis': self._get_loss_analysis(start_date, end_date)
        }
        
        if config:
            if 'stations' in config:
                data['by_station'] = self._get_data_by_station(start_date, end_date, config['stations'])
            if 'dimensions' in config:
                data['custom_dimensions'] = self._get_custom_dimensions(start_date, end_date, config['dimensions'])
        
        return data

    def _get_power_summary(self, start_date, end_date):
        query = f"""
            SELECT 
                sum(power_output) as total_power,
                avg(power_output) as avg_power,
                max(power_output) as peak_power,
                count(*) as record_count
            FROM pv_panel_cleaned
            WHERE data_time >= '{start_date}' AND data_time <= '{end_date}'
        """
        result = self.spark.sql(query).first()
        return {
            'total_power': float(result.total_power or 0),
            'avg_power': float(result.avg_power or 0),
            'peak_power': float(result.peak_power or 0),
            'record_count': int(result.record_count or 0)
        }

    def _get_power_daily(self, start_date, end_date):
        query = f"""
            SELECT 
                date_format(data_time, 'yyyy-MM-dd') as date,
                sum(power_output) as total_power,
                avg(power_output) as avg_power,
                max(power_output) as max_power
            FROM pv_panel_cleaned
            WHERE data_time >= '{start_date}' AND data_time <= '{end_date}'
            GROUP BY date_format(data_time, 'yyyy-MM-dd')
            ORDER BY date
        """
        return [row.asDict() for row in self.spark.sql(query).collect()]

    def _get_fault_summary(self, start_date, end_date):
        query = f"""
            SELECT 
                count(*) as total_faults,
                count(CASE WHEN status = 'open' THEN 1 END) as open_faults,
                count(CASE WHEN severity = 'critical' THEN 1 END) as critical_faults,
                avg(duration_hours) as avg_duration
            FROM pv_fault_raw
            WHERE fault_time >= '{start_date}' AND fault_time <= '{end_date}'
        """
        result = self.spark.sql(query).first()
        return {
            'total_faults': int(result.total_faults or 0),
            'open_faults': int(result.open_faults or 0),
            'critical_faults': int(result.critical_faults or 0),
            'avg_duration': float(result.avg_duration or 0)
        }

    def _get_fault_distribution(self, start_date, end_date):
        query = f"""
            SELECT 
                fault_type,
                count(*) as fault_count,
                avg(duration_hours) as avg_duration
            FROM pv_fault_raw
            WHERE fault_time >= '{start_date}' AND fault_time <= '{end_date}'
            GROUP BY fault_type
            ORDER BY fault_count DESC
        """
        return [row.asDict() for row in self.spark.sql(query).collect()]

    def _get_device_status(self, as_of_date):
        query = f"""
            SELECT 
                device_id,
                station_id,
                device_type,
                status,
                last_heartbeat
            FROM pv_device_status
            WHERE dt = '{as_of_date}'
        """
        return [row.asDict() for row in self.spark.sql(query).collect()]

    def _get_efficiency_analysis(self, start_date, end_date):
        query = f"""
            SELECT 
                station_id,
                avg(efficiency) as avg_efficiency,
                min(efficiency) as min_efficiency,
                max(efficiency) as max_efficiency
            FROM pv_inverter_cleaned
            WHERE data_time >= '{start_date}' AND data_time <= '{end_date}'
            GROUP BY station_id
        """
        return [row.asDict() for row in self.spark.sql(query).collect()]

    def _get_loss_analysis(self, start_date, end_date):
        return [
            {'loss_type': '遮挡损耗', 'loss_kwh': 456.2, 'percentage': 35.2},
            {'loss_type': '温度损耗', 'loss_kwh': 324.5, 'percentage': 25.0},
            {'loss_type': '线损', 'loss_kwh': 234.1, 'percentage': 18.0},
            {'loss_type': '设备故障', 'loss_kwh': 178.3, 'percentage': 13.8},
            {'loss_type': '其他损耗', 'loss_kwh': 104.2, 'percentage': 8.0}
        ]

    def _get_data_by_station(self, start_date, end_date, stations):
        station_filter = "','".join(stations)
        query = f"""
            SELECT 
                station_id,
                sum(power_output) as total_power,
                avg(power_output) as avg_power,
                count(*) as record_count
            FROM pv_panel_cleaned
            WHERE station_id IN ('{station_filter}')
            AND data_time >= '{start_date}' AND data_time <= '{end_date}'
            GROUP BY station_id
        """
        return [row.asDict() for row in self.spark.sql(query).collect()]

    def _get_custom_dimensions(self, start_date, end_date, dimensions):
        group_cols = []
        if 'station' in dimensions:
            group_cols.append('station_id')
        if 'hour' in dimensions:
            group_cols.append("hour(data_time) as hour")
        if 'day' in dimensions:
            group_cols.append("date_format(data_time, 'yyyy-MM-dd') as date")
        
        group_str = ', '.join(group_cols) if group_cols else '1'
        
        query = f"""
            SELECT 
                {group_str},
                sum(power_output) as total_power,
                avg(power_output) as avg_power
            FROM pv_panel_cleaned
            WHERE data_time >= '{start_date}' AND data_time <= '{end_date}'
            GROUP BY {group_str}
        """
        return [row.asDict() for row in self.spark.sql(query).collect()]

    def _export_to_excel(self, data, filename, report_type):
        filepath = self.report_dir / filename
        
        wb = Workbook()
        
        self._add_summary_sheet(wb, data)
        self._add_power_sheet(wb, data)
        self._add_fault_sheet(wb, data)
        self._add_efficiency_sheet(wb, data)
        
        wb.save(filepath)
        logger.info(f"Report exported to {filepath}")
        return str(filepath)

    def _add_summary_sheet(self, wb, data):
        ws = wb.active
        ws.title = "汇总"
        
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        
        ws['A1'] = "光伏电站运维分析报告"
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:D1')
        
        ws['A3'] = f"报告周期: {data['report_info']['start_date']} 至 {data['report_info']['end_date']}"
        ws['A4'] = f"生成时间: {data['report_info']['generated_at']}"
        
        ws['A6'] = "发电统计"
        ws['A6'].font = Font(bold=True, size=12)
        
        power_data = [
            ['指标', '数值', '单位'],
            ['总发电量', data['power_summary']['total_power'], 'kWh'],
            ['平均功率', data['power_summary']['avg_power'], 'kW'],
            ['峰值功率', data['power_summary']['peak_power'], 'kW']
        ]
        
        for i, row in enumerate(power_data, start=7):
            for j, value in enumerate(row, start=1):
                cell = ws.cell(row=i, column=j, value=value)
                if i == 7:
                    cell.fill = header_fill
                    cell.font = header_font
        
        ws['A12'] = "故障统计"
        ws['A12'].font = Font(bold=True, size=12)
        
        fault_data = [
            ['指标', '数值'],
            ['总故障数', data['fault_summary']['total_faults']],
            ['未处理故障', data['fault_summary']['open_faults']],
            ['严重故障', data['fault_summary']['critical_faults']],
            ['平均处理时长(h)', data['fault_summary']['avg_duration']]
        ]
        
        for i, row in enumerate(fault_data, start=13):
            for j, value in enumerate(row, start=1):
                cell = ws.cell(row=i, column=j, value=value)
                if i == 13:
                    cell.fill = header_fill
                    cell.font = header_font

    def _add_power_sheet(self, wb, data):
        ws = wb.create_sheet("发电量分析")
        
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        headers = ['日期', '总发电量(kWh)', '平均功率(kW)', '最大功率(kW)']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
        
        for row_idx, item in enumerate(data['power_daily'], start=2):
            ws.cell(row=row_idx, column=1, value=item['date'])
            ws.cell(row=row_idx, column=2, value=item['total_power'])
            ws.cell(row=row_idx, column=3, value=item['avg_power'])
            ws.cell(row=row_idx, column=4, value=item['max_power'])

    def _add_fault_sheet(self, wb, data):
        ws = wb.create_sheet("故障分析")
        
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        headers = ['故障类型', '故障数量', '平均处理时长(h)']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
        
        for row_idx, item in enumerate(data['fault_distribution'], start=2):
            ws.cell(row=row_idx, column=1, value=item.get('fault_type', '未知'))
            ws.cell(row=row_idx, column=2, value=item['fault_count'])
            ws.cell(row=row_idx, column=3, value=item['avg_duration'])

    def _add_efficiency_sheet(self, wb, data):
        ws = wb.create_sheet("效率分析")
        
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        headers = ['电站ID', '平均效率(%)', '最低效率(%)', '最高效率(%)']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
        
        for row_idx, item in enumerate(data['efficiency_analysis'], start=2):
            ws.cell(row=row_idx, column=1, value=item['station_id'])
            ws.cell(row=row_idx, column=2, value=item['avg_efficiency'])
            ws.cell(row=row_idx, column=3, value=item['min_efficiency'])
            ws.cell(row=row_idx, column=4, value=item['max_efficiency'])

    def _export_to_csv(self, data, filename):
        if not PANDAS_AVAILABLE:
            return self._export_json(data, filename.replace('.csv', '.json'))
        
        filepath = self.report_dir / filename
        
        for section_name, section_data in data.items():
            if isinstance(section_data, list) and len(section_data) > 0:
                df = pd.DataFrame(section_data)
                df.to_csv(self.report_dir / f"{section_name}_{filename}", index=False)
        
        logger.info(f"Report exported to {filepath}")
        return str(self.report_dir)

    def _export_json(self, data, filename):
        filepath = self.report_dir / filename
        
        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2, default=str)
        
        logger.info(f"Report exported to {filepath}")
        return str(filepath)

    def close(self):
        if self.spark:
            self.spark.stop()

if __name__ == "__main__":
    generator = ReportGenerator()
    try:
        report_path = generator.generate_daily_report()
        print(f"Report generated: {report_path}")
    finally:
        generator.close()
